"""
Railway Server - No Scrapers
Only serves the API and frontend, reads from prices.json and prices_ftn.json.gz
"""
import uvicorn
import json
import os
import gzip
import re
import sys
from fastapi import FastAPI, Request, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import Optional, List
import subprocess
from starlette.middleware.base import BaseHTTPMiddleware

# Fix encoding for Windows
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# ==========================================
# ⚙️ CONFIGURATION
# ==========================================
DATA_FILE_VIAGOGO = 'prices.json'
DATA_FILE_FTN = 'prices_ftn.json.gz'
DATA_FILE_FTN_LEGACY = 'prices_ftn.json'
GAMES_FILE = 'all_games_to_scrape.json'
TEAMS_DATA_FILE = 'ftn_teams_data.json'
# Railway sets PORT dynamically - use whatever Railway provides
# Railway will set PORT environment variable automatically
# Railway will set PORT environment variable automatically
from dotenv import load_dotenv
load_dotenv()

PORT = int(os.environ.get('PORT', '8000'))  # Railway always sets PORT, but keep default for local dev

# Gemini API Key (User Provided)
# Gemini API Key (User Provided)
# os.environ['GEMINI_API_KEY'] = '...' # Key removed for security. Please set GEMINI_API_KEY env var.

# ==========================================
# FastAPI App
# ==========================================
from contextlib import asynccontextmanager

# Lifespan event handler (replaces deprecated on_event)
@asynccontextmanager
async def lifespan(app_instance: FastAPI):
    # Startup - Railway handles port/IP configuration
    print('[STARTUP] FastAPI application started', flush=True)
    try:
        yield
    finally:
        # Shutdown (if needed)
        print('[SHUTDOWN] FastAPI application shutting down', flush=True)

app = FastAPI(title="Viagogo Monitor API", lifespan=lifespan)

# Request logging middleware
class LoggingMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        print(f'[REQUEST] {request.method} {request.url.path}')
        response = await call_next(request)
        print(f'[RESPONSE] {request.method} {request.url.path} -> {response.status_code}')
        return response

app.add_middleware(LoggingMiddleware)

app.add_middleware(
    CORSMiddleware,
    allow_origins=['*'],
    allow_credentials=True,
    allow_methods=['*'],
    allow_headers=['*'],
)

# ---------------------------------------------------------
# JSON Data Utility
# ---------------------------------------------------------
def load_data(file_path):
    if not os.path.exists(file_path): 
        print(f'[WARN] File not found: {file_path}')
        return []
    try:
        if file_path.endswith('.gz'):
            with gzip.open(file_path, 'rt', encoding='utf-8') as f:
                data = json.load(f)
        else:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            print(f'[INFO] Loaded {len(data)} records from {file_path}')
            return data
    except Exception as e:
        print(f'[ERROR] Failed to load {file_path}: {e}')
        return []

# ---------------------------------------------------------
# API Endpoints
# ---------------------------------------------------------
@app.get('/matches')
def get_matches():
    try:
        print('[API] /matches endpoint called')
        if not os.path.exists(GAMES_FILE): 
            print(f'[WARN] {GAMES_FILE} not found')
            return []
        with open(GAMES_FILE, 'r', encoding='utf-8') as f: 
            games = json.load(f)
        
        def get_match_number(match_name):
            m = re.search(r'Match (\d+)', match_name)
            return int(m.group(1)) if m else 9999

        matches_list = [{'match_name': g['match_name'], 'match_url': g['url']} for g in games]
        matches_list.sort(key=lambda x: get_match_number(x['match_name']))
        print(f'[API] Returning {len(matches_list)} matches')
        return matches_list
    except Exception as e:
        print(f'[ERROR] API Error: {e}')
        import traceback
        traceback.print_exc()
        return []

@app.get('/history')
def get_history(match_url: str):
    try:
        print(f"[API] History Request for URL: {match_url[:50]}...")
        
        # 1. LOAD VIAGOGO DATA
        viagogo_data = load_data(DATA_FILE_VIAGOGO)
        v_match_data = []
        
        # Extract ID from requested URL
        # e.g. .../World-Cup-Tickets/E-153033506?Currency=... -> E-153033506
        req_id = None
        match = re.search(r'/(E-\d+)', match_url)
        if match: 
            req_id = match.group(1)
        
        print(f"[API] Looking for Viagogo ID: {req_id}")

        for row in viagogo_data:
            # Extract ID from stored URL (remove query params for comparison)
            stored_url = row.get('match_url', '').split('?')[0].split('&')[0]
            stored_id = None
            m_stored = re.search(r'/(E-\d+)', stored_url)
            if m_stored: 
                stored_id = m_stored.group(1)
            
            # Clean request URL too
            clean_req_url = match_url.split('?')[0].split('&')[0]
            
            # Match by ID if possible (most reliable)
            if req_id and stored_id:
                if req_id == stored_id:
                    v_match_data.append(row)
                    continue
            
            # Match by URL (exact or partial)
            if clean_req_url in stored_url or stored_url in clean_req_url:
                v_match_data.append(row)
                continue
            
            # Match by match number in match_name
            match_name = row.get('match_name', '')
            match_num = re.search(r'Match (\d+)', match_name)
            if match_num:
                url_match_num = re.search(r'Match (\d+)', match_url, re.IGNORECASE)
                if url_match_num and match_num.group(1) == url_match_num.group(1):
                    v_match_data.append(row)
                    continue
                 
        print(f"[API] Found {len(v_match_data)} Viagogo records.")
        v_match_data.sort(key=lambda x: x.get('timestamp', ''))

        # 2. IDENTIFY MATCH FOR FTN
        # Prefer gz if present; fall back to legacy json for first-time migration.
        ftn_data_file = DATA_FILE_FTN if os.path.exists(DATA_FILE_FTN) else DATA_FILE_FTN_LEGACY
        ftn_data = load_data(ftn_data_file)
        f_match_data = []
        
        match_number = None
        m = re.search(r'Match (\d+)', match_url, re.IGNORECASE)
        
        # New fallback: Look in GAMES_FILE if URL does not have match number
        if not m and os.path.exists(GAMES_FILE):
             try:
                 with open(GAMES_FILE, 'r', encoding='utf-8') as f: 
                     games = json.load(f)
                 # Find game with this URL (ignoring query params)
                 clean_input_url = match_url.split('?')[0]
                 
                 for g in games:
                     if g['url'].split('?')[0] == clean_input_url:
                         m = re.search(r'Match (\d+)', g['match_name'], re.IGNORECASE)
                         break
             except Exception as e:
                 print(f'[ERROR] Error loading games file: {e}')

        if not m and v_match_data:
             m = re.search(r'Match (\d+)', v_match_data[0].get('match_name', ''), re.IGNORECASE)
        
        if m:
            match_number = m.group(1)
            # Find FTN records for this Match #
            f_match_data = [d for d in ftn_data if f'Match {match_number}' in d.get('match_name', '')]
            f_match_data.sort(key=lambda x: x.get('timestamp', ''))
            print(f"[API] Found {len(f_match_data)} FTN records for Match {match_number}")
        
        def process_source_data(data_list):
            if not data_list: 
                return {}, []
            categories = sorted(list(set(d.get('category', '') for d in data_list if d.get('category'))))
            result_data = {}
            for cat in categories:
                cat_rows = [d for d in data_list if d.get('category') == cat]
                result_data[cat] = [{'timestamp': d.get('timestamp', ''), 'price': d.get('price', 0)} for d in cat_rows]
            return result_data, categories

        v_processed, v_cats = process_source_data(v_match_data)
        f_processed, f_cats = process_source_data(f_match_data)

        result = {
            'viagogo': {'categories': v_cats, 'data': v_processed},
            'ftn': {'categories': f_cats, 'data': f_processed},
            'currency': 'USD'
        }
        
        print(f"[API] Returning: Viagogo categories: {v_cats}, FTN categories: {f_cats}")
        return result

    except Exception as e:
        print(f'[ERROR] History Error: {e}')
        import traceback
        traceback.print_exc()
        return {'viagogo': {'categories': [], 'data': {}}, 'ftn': {'categories': [], 'data': {}}}

# ---------------------------------------------------------
# Auto-Build Frontend
# ---------------------------------------------------------
client_dist = 'frontend/dist'

def build_frontend():
    print('[INFO] Building frontend...')
    try:
        if not os.path.exists('frontend'):
            print('[ERROR] Frontend folder missing!')
            return False
            
        npm_cmd = 'npm.cmd' if os.name == 'nt' else 'npm'
        
        # Install dependencies
        print('[INFO] Installing npm dependencies...')
        result = subprocess.run(
            [npm_cmd, 'install'], 
            cwd='frontend', 
            shell=True,
            capture_output=True,
            text=True
        )
        if result.returncode != 0:
            print(f'[WARN] npm install had issues: {result.stderr[:200]}')
        
        # Build
        print('[INFO] Building frontend...')
        result = subprocess.run(
            [npm_cmd, 'run', 'build'], 
            cwd='frontend', 
            shell=True,
            capture_output=True,
            text=True
        )
        if result.returncode == 0:
            print('[OK] Frontend build complete!')
            return True
        else:
            print(f'[ERROR] Build failed: {result.stderr[:500]}')
            return False
    except Exception as e:
        print(f'[ERROR] Build Failed: {e}')
        import traceback
        traceback.print_exc()
        return False

# Build frontend if needed
if not os.path.exists(f'{client_dist}/index.html'):
    print('[WARN] Frontend build not found. Building now...')
    if not build_frontend():
        print('[ERROR] Frontend build failed! Dashboard may not work.')
else:
    print('[OK] Frontend build found.')
    # Verify it's a valid build
    if not os.path.exists(f'{client_dist}/index.html'):
        print('[ERROR] Frontend dist/index.html missing!')
    else:
        print(f'[OK] Frontend index.html found at {client_dist}/index.html')

# Mount static files - MUST be done before catch-all route
if os.path.exists(client_dist):
    # Mount assets directory (JS, CSS files)
    assets_path = f'{client_dist}/assets'
    if os.path.exists(assets_path):
        try:
            app.mount('/assets', StaticFiles(directory=assets_path, html=False), name='assets')
            # List files in assets for debugging
            asset_files = os.listdir(assets_path)
            print(f'[OK] Static assets mounted at /assets ({len(asset_files)} files)')
            for f in asset_files[:5]:  # Show first 5 files
                print(f'      - {f}')
        except Exception as e:
            print(f'[WARN] Failed to mount assets: {e}')
            import traceback
            traceback.print_exc()
    else:
        print(f'[WARN] Assets directory not found: {assets_path}')
        print(f'[INFO] Contents of {client_dist}: {os.listdir(client_dist) if os.path.exists(client_dist) else "N/A"}')
    
    print('[OK] Static file routes configured')
else:
    print(f'[ERROR] Frontend dist directory not found: {client_dist}')

def get_team_currency(team_key):
    """Get currency for a team from teams_list.json"""
    try:
        teams_list_file = 'teams_list.json'
        if os.path.exists(teams_list_file):
            with open(teams_list_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                for team in data.get('teams', []):
                    if team.get('team_key') == team_key:
                        return team.get('currency', 'USD')
    except Exception as e:
        print(f'[WARN] Could not load currency for {team_key}: {e}')
    
    # Default currency mapping based on team name/country
    default_currencies = {
        'arsenal': 'GBP',
        'barcelona': 'EUR',
        'real-madrid': 'EUR',
        'manchester': 'GBP',
        'liverpool': 'GBP',
        'chelsea': 'GBP',
        'tottenham': 'GBP',
    }
    
    # Check if team_key matches any default
    for key, currency in default_currencies.items():
        if key in team_key.lower():
            return currency
    
    return 'USD'  # Default

@app.get('/teams')
def get_teams():
    """Get list of available teams"""
    try:
        if not os.path.exists(TEAMS_DATA_FILE):
            print(f'[INFO] /teams: {TEAMS_DATA_FILE} not found, returning empty list')
            return []
        
        # Try to load JSON, handle merge conflicts
        try:
            with open(TEAMS_DATA_FILE, 'r', encoding='utf-8') as f:
                content = f.read()
                # Check for merge conflicts
                if '<<<<<<<' in content or '=======' in content or '>>>>>>>' in content:
                    print(f'[WARN] /teams: {TEAMS_DATA_FILE} contains merge conflicts, returning empty list')
                    return []
                data = json.loads(content)
        except json.JSONDecodeError as json_err:
            print(f'[ERROR] /teams: JSON decode error: {json_err}')
            return []
        except Exception as file_err:
            print(f'[ERROR] /teams: File read error: {file_err}')
            return []
        
        if not isinstance(data, dict):
            print(f'[ERROR] /teams: Expected dict, got {type(data)}')
            return []
        
        teams = []
        for team_key, team_data in data.items():
            if not isinstance(team_data, dict):
                continue
            try:
                currency = get_team_currency(team_key)
                teams.append({
                    'key': str(team_key),
                    'name': str(team_data.get('team_name', team_key.title())),
                    'url': str(team_data.get('team_url', '')),
                    'last_updated': team_data.get('last_updated'),
                    'game_count': len(team_data.get('games', [])) if isinstance(team_data.get('games'), list) else 0,
                    'currency': str(currency)
                })
            except Exception as team_err:
                print(f'[WARN] /teams: Error processing team {team_key}: {team_err}')
                continue
        
        print(f'[INFO] /teams: Returning {len(teams)} teams')
        return teams
    except Exception as e:
        print(f'[ERROR] /teams error: {e}')
        import traceback
        traceback.print_exc()
        return []

def parse_game_date(date_str):
    """Parse date string in format DD/MM/YY to datetime"""
    try:
        if not date_str:
            return None
        # Format: "27/12/25" -> DD/MM/YY
        parts = date_str.split('/')
        if len(parts) == 3:
            day, month, year = int(parts[0]), int(parts[1]), int(parts[2])
            # Convert 2-digit year to 4-digit (assuming 20xx for years < 50, 19xx otherwise)
            if year < 50:
                year += 2000
            else:
                year += 1900
            from datetime import datetime
            return datetime(year, month, day)
    except Exception as e:
        print(f'[WARN] Could not parse date {date_str}: {e}')
    return None

@app.get('/teams/{team_key}')
def get_team_games(team_key: str):
    """Get all games for a specific team (filter out past games)"""
    try:
        if not os.path.exists(TEAMS_DATA_FILE):
            return []
        with open(TEAMS_DATA_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
        if team_key not in data:
            return []
        team_data = data[team_key]
        games = []
        from datetime import datetime
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        
        for game in team_data.get('games', []):
            game_date = parse_game_date(game.get('date'))
            # Filter out past games (games before today)
            if game_date and game_date < today:
                continue  # Skip past games
            
            games.append({
                'url': game.get('url'),
                'match_name': game.get('match_name'),
                'opponent': game.get('opponent'),
                'date': game.get('date'),
                'latest_prices': game.get('latest_prices', {}),
                'last_scraped': game.get('last_scraped'),
                'price_history_count': len(game.get('price_history', []))
            })
        return games
    except Exception as e:
        print(f'[ERROR] /teams/{team_key} error: {e}')
        return []

@app.get('/teams/{team_key}/game/{game_index}')
def get_game_prices(team_key: str, game_index: int):
    """Get price history for a specific game"""
    try:
        if not os.path.exists(TEAMS_DATA_FILE):
            return {'prices': [], 'game': None, 'currency': 'USD'}
        with open(TEAMS_DATA_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
        if team_key not in data:
            return {'prices': [], 'game': None, 'currency': 'USD'}
        
        # Filter out past games when getting the list
        from datetime import datetime
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        all_games = data[team_key].get('games', [])
        future_games = []
        for game in all_games:
            game_date = parse_game_date(game.get('date'))
            if game_date and game_date >= today:
                future_games.append(game)
        
        if game_index < 0 or game_index >= len(future_games):
            return {'prices': [], 'game': None, 'currency': 'USD'}
        
        game = future_games[game_index]
        currency = get_team_currency(team_key)
        return {
            'game': {
                'match_name': game.get('match_name'),
                'url': game.get('url'),
                'opponent': game.get('opponent'),
                'date': game.get('date')
            },
            'prices': game.get('price_history', []),
            'latest_prices': game.get('latest_prices', {}),
            'currency': currency
        }
    except Exception as e:
        print(f'[ERROR] /teams/{team_key}/game/{game_index} error: {e}')
        return {'prices': [], 'game': None, 'currency': 'USD'}

@app.get('/vite.svg')
async def serve_vite_svg():
    vite_svg_path = f'{client_dist}/vite.svg'
    if os.path.exists(vite_svg_path):
        return FileResponse(vite_svg_path, media_type='image/svg+xml')
    return {'error': 'vite.svg not found'}

# Test endpoint to verify server is working - lightweight, no data loading
# ==========================================
# Ticket Offers Manager API
# ==========================================
from tickets_storage import (
    load_sellers, create_seller, get_seller_by_name,
    save_offer, get_offer_by_id, load_offers_from_file, rebuild_index,
    load_index, BASE_DIR
)
from tickets_parser import parse_ticket_message, interpret_search_query
from pathlib import Path
from datetime import datetime, timedelta

# Pydantic models for request/response
class CreateSellerRequest(BaseModel):
    name: str

class AskRequest(BaseModel):
    query: str

class ParseOfferRequest(BaseModel):
    seller: str
    raw: str
    use_ai: bool = True

class SaveOfferRequest(BaseModel):
    seller: str
    raw: str
    parsed: dict

@app.get('/api/tickets/sellers')
def get_sellers():
    """Return sellers list"""
    try:
        sellers_data = load_sellers()
        return sellers_data.get('sellers', [])
    except Exception as e:
        print(f'[ERROR] /api/tickets/sellers: {e}')
        return []

@app.post('/api/tickets/ask')
def ask_tickets(request: AskRequest):
    """Ask AI for tickets (NL Search)"""
    try:
        api_key = os.environ.get("GEMINI_API_KEY")
        if not api_key:
             raise HTTPException(status_code=500, detail="AI not configured")
             
        filters = interpret_search_query(request.query, api_key)
        if not filters:
             return []
             
        print(f"[INFO] AI Filters: {filters}")
        
        # Map filters to search logic
        offers = _search_offers_logic(
            match=filters.get('match'),
            seller=filters.get('seller'),
            category=filters.get('category'),
            min_price=filters.get('min_price'),
            max_price=filters.get('max_price'),
            min_quantity=filters.get('min_quantity'),
            max_quantity=None,
            keyword=filters.get('keyword'),
            range='all'
        )
        
        # Apply Sorting
        sort_by = filters.get('sort_by')
        if sort_by == 'price_asc':
            def get_min_price(o):
                 prices = [float(l.get('price', 999999)) for l in o.get('lines', []) if l.get('price')]
                 return min(prices) if prices else 999999
            offers.sort(key=get_min_price)
            
        elif sort_by == 'price_desc':
            def get_max_price(o):
                 prices = [float(l.get('price', 0)) for l in o.get('lines', []) if l.get('price')]
                 return max(prices) if prices else 0
            offers.sort(key=get_max_price, reverse=True)
            
        elif sort_by == 'date_desc':
            offers.sort(key=lambda x: x.get('created_at', ''), reverse=True)
            
        # Limit
        limit = filters.get('limit', 50)
        return offers[:limit]
        
    except Exception as e:
        print(f'[ERROR] /api/tickets/ask: {e}')
        import traceback
        traceback.print_exc()
        return []

@app.post('/api/tickets/sellers')
def create_seller_endpoint(request: CreateSellerRequest):
    """Create seller with case-insensitive dedupe"""
    try:
        seller = create_seller(request.name)
        return seller
    except Exception as e:
        print(f'[ERROR] /api/tickets/sellers POST: {e}')
        raise HTTPException(status_code=500, detail=str(e))

@app.post('/api/tickets/offers/parse')
def parse_offer(request: ParseOfferRequest):
    """Parse raw text into structured data without saving"""
    try:
        print(f'[INFO] /api/tickets/offers/parse called with seller: {request.seller}, raw length: {len(request.raw) if request.raw else 0}')
        if not request.raw:
            raise HTTPException(status_code=400, detail='Raw text is required')
            
        # If use_ai is True, pass None to use Env Var. If False, pass "" to disable AI.
        key_arg = None if request.use_ai else ""
        parsed = parse_ticket_message(request.raw, api_key=key_arg)
        
        print(f'[INFO] Parse result: match={parsed.get("match")}, lines={len(parsed.get("lines", []))}, status={parsed.get("parse_status")}')
        return parsed
    except HTTPException:
        raise
    except Exception as e:
        print(f'[ERROR] /api/tickets/offers/parse: {e}')
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.post('/api/tickets/offers')
def save_offer_endpoint(request: SaveOfferRequest):
    """Save offer to weekly JSONL and update index"""
    try:
        # Get or create seller
        seller_obj = get_seller_by_name(request.seller)
        if not seller_obj:
            seller_obj = create_seller(request.seller)
        
        # Create offer record
        now = datetime.utcnow()
        offer = {
            'seller': request.seller,
            'seller_id': seller_obj['id'],
            'match': request.parsed.get('match'),
            'event': request.parsed.get('event'),
            'lines': request.parsed.get('lines', []),
            'raw': request.raw,
            'parse_status': request.parsed.get('parse_status', 'ok'),
            'warnings': request.parsed.get('warnings', []),
            'created_at': now.isoformat()
        }
        
        # Save offer
        success = save_offer(offer)
        if not success:
            raise HTTPException(status_code=500, detail='Failed to save offer')
        
        return {'success': True, 'id': offer.get('id')}
    except Exception as e:
        print(f'[ERROR] /api/tickets/offers POST: {e}')
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.get('/api/tickets/offers/search')
def search_offers(
    match: Optional[int] = None,
    seller: Optional[str] = None,
    category: Optional[str] = None,
    min_price: Optional[float] = None,
    max_price: Optional[float] = None,
    min_quantity: Optional[int] = None,
    max_quantity: Optional[int] = None,
    keyword: Optional[str] = None,
    range: str = 'all'
):
    """Search offers using index.json and fallback scanning"""
    return _search_offers_logic(match, seller, category, min_price, max_price, min_quantity, max_quantity, keyword, range)

def _search_offers_logic(match, seller, category, min_price, max_price, min_quantity, max_quantity, keyword, range):
    try:
        # Debugging
        print(f"[SEARCH] Filters: match={match}, seller={seller}, cat={category}, price={min_price}-{max_price}", flush=True)
        
        index_data = load_index()
        # Ensure we have valid sets
        if 'match' not in index_data: index_data['match'] = {}
        if 'seller' not in index_data: index_data['seller'] = {}
        if 'category' not in index_data: index_data['category'] = {}
        if 'offers_by_id' not in index_data: index_data['offers_by_id'] = {}

        candidate_ids = None
        
        # Time range filter
        now = datetime.utcnow()
        if range == '7d':
            cutoff = now - timedelta(days=7)
        elif range == '30d':
            cutoff = now - timedelta(days=30)
        else:
            cutoff = None
        
        # --- INDEX FILTERING ---
        used_index = False
        
        # Match Filter
        if match is not None:
            used_index = True
            match_key = str(match)
            # If match key exists, get IDs. If not, we get empty set, which is correct (no results).
            # But we must initialize candidate_ids.
            ids = set(index_data['match'].get(match_key, []))
            if candidate_ids is None:
                candidate_ids = ids
            else:
                candidate_ids &= ids

        # Seller Filter
        if seller:
            used_index = True
            seller_obj = get_seller_by_name(seller)
            if seller_obj:
                seller_id = seller_obj['id']
                ids = set(index_data['seller'].get(seller_id, []))
                if candidate_ids is None:
                    candidate_ids = ids
                else:
                    candidate_ids &= ids
            else:
                # Seller not found -> no results
                candidate_ids = set()
        
        # Category Filter (Skipping index for fuzzy matching capabilities)
        
        # --- FALLBACK ---
        # If candidate_ids IS NONE, it means no index filters were applied -> Search All.
        # If candidate_ids IS EMPTY SET (not None), it means index filters returned 0 results.
        
        if candidate_ids is None:
            # Fallback to all indexed offers
            candidate_ids = set(index_data['offers_by_id'].keys())
            
            # If index is empty/broken, try file scan?
            if not candidate_ids:
                 offer_files = list(BASE_DIR.glob('offers_*.jsonl'))
                 for file_path in offer_files:
                     offers = load_offers_from_file(file_path)
                     for offer in offers:
                         candidate_ids.add(offer.get('id'))

        if not candidate_ids:
             print("[SEARCH] No candidates found.", flush=True)
             return []

        # --- LOAD & FILTER ---
        results = []
        offers_by_id = index_data.get('offers_by_id', {})
        
        # Group by file to minimize IO
        file_groups = {}
        for oid in candidate_ids:
            if oid in offers_by_id:
                fname = offers_by_id[oid]['file']
                if fname not in file_groups: file_groups[fname] = []
                file_groups[fname].append(oid)
        
        print(f"[SEARCH] Scanning {len(file_groups)} files...", flush=True)

        for fname, oids in file_groups.items():
            file_path = BASE_DIR / fname
            if not file_path.exists(): continue
            
            file_offers = load_offers_from_file(file_path)
            target_ids = set(oids)
            
            for offer in file_offers:
                if offer.get('id') not in target_ids:
                    continue
                
                # Time Check
                if 'created_at' in offer:
                    created_at = datetime.fromisoformat(offer['created_at'])
                else:
                    created_at = now
                if cutoff and created_at < cutoff:
                    continue
                
                # Top Level Filters (Seller, Keyword)
                if seller and str(offer.get('seller', '')).lower() != str(seller).lower():
                    continue
                if keyword and keyword.lower() not in offer.get('raw', '').lower():
                    continue

                # Filter Lines to show specific matches users asked for
                lines = offer.get('lines', [])
                valid_lines = []
                
                # Check Global Match
                global_match_satisfies = False
                if match is not None:
                     if str(offer.get('match')) == str(match):
                         global_match_satisfies = True
                
                # Determine filters
                has_filters = (match is not None or category or min_price is not None or 
                               max_price is not None or min_quantity is not None or max_quantity is not None)
                
                if not has_filters:
                    results.append(offer)
                    continue

                for line in lines:
                    # Match Check
                    if match is not None:
                        l_match = line.get('match')
                        o_match = offer.get('match')
                        # Use line match if present, else fallback to offer match
                        effective_match = l_match if l_match is not None else o_match
                        
                        if effective_match is None or str(effective_match) != str(match):
                            continue
                    # Category
                    if category:
                        l_cat = str(line.get('category', '')).lower()
                        t_cat = str(category).lower()
                        if t_cat not in l_cat and l_cat not in t_cat:
                            continue
                    
                    # Price
                    price = float(line.get('price', 0))
                    if min_price is not None and price < min_price: continue
                    if max_price is not None and price > max_price: continue
                    
                    # Qty
                    qty = line.get('quantity')
                    if min_quantity is not None and (qty is None or qty < min_quantity): continue
                    if max_quantity is not None and (qty is None or qty > max_quantity): continue
                    
                    valid_lines.append(line)
                
                if valid_lines:
                    # Return copy with only valid lines
                    offer_copy = offer.copy()
                    offer_copy['lines'] = valid_lines
                    
                    # Optimization: If all lines are for the same match, promote it to top-level for UI clarity
                    if offer_copy.get('match') is None:
                         unique_matches = set(str(l.get('match')) for l in valid_lines if l.get('match'))
                         if len(unique_matches) == 1:
                             try:
                                offer_copy['match'] = int(list(unique_matches)[0])
                             except: pass

                    results.append(offer_copy)

        print(f"[SEARCH] Found {len(results)} results.", flush=True)
        return results

    except Exception as e:
        print(f'[ERROR] Search Logic: {e}')
        # import traceback
        # traceback.print_exc()
        return []
        traceback.print_exc()
        return []

@app.get('/api/tickets/export')
def export_tickets(
    match: Optional[int] = None,
    seller: Optional[str] = None,
    category: Optional[str] = None,
    range: str = 'all'
):
    """Export tickets to Excel"""
    try:
        import pandas as pd
        import io
        from fastapi.responses import StreamingResponse
        
        # Refactor search logic to reuse it or just duplicate simple logic
        # We'll use the search logic function extracted above
        offers = _search_offers_logic(match, seller, category, None, None, None, None, None, range)
        
        # Flatten data
        rows = []
        for offer in offers:
            seller_name = offer.get('seller', 'Unknown')
            offer_date = offer.get('created_at', '')
            top_match = offer.get('match')
            event = offer.get('event')
            
            for line in offer.get('lines', []):
                # Determine match for this line
                line_match = line.get('match')
                final_match = line_match if line_match else top_match
                
                # Strict Filtering for Export
                if match is not None:
                    try:
                        # Allow robust comparison (int/str)
                        if final_match is None or int(final_match) != int(match):
                            continue
                    except (ValueError, TypeError):
                        # If match is not a valid number (e.g. "None" or text), skip if we needed a specific match
                         continue
                
                if category:
                   if str(line.get('category')) != str(category):
                       continue

                rows.append({
                    'Match': final_match,
                    'Event': event,
                    'Date': offer_date,
                    'Seller': seller_name,
                    'Category': line.get('category'),
                    'Quantity': line.get('quantity'),
                    'Cost Price': line.get('price'),
                    'My Price': line.get('my_price'),
                    'Currency': line.get('currency'),
                    'Raw': offer.get('raw')[:50] + '...' # Truncate raw
                })
        
        if not rows:
             return {'error': 'No data found to export'}

        df = pd.DataFrame(rows)
        
        # Grouping/Breaking Logic
        # "Group by Match, Break by Seller"
        # Since Excel is a flat format (unless multiple sheets), we will sort.
        # Primary Sort: Match
        # Secondary Sort: Seller
        
        # Ensure Match is numeric for sorting if possible
        df['Match'] = pd.to_numeric(df['Match'], errors='coerce').fillna(999999)
        
        df = df.sort_values(by=['Match', 'Seller'])
        
        # Create Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Tickets')
            
            # Application of styles using openpyxl
            workbook = writer.book
            worksheet = writer.sheets['Tickets']
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Header Style
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='2b5797', end_color='2b5797', fill_type='solid') # Nice Blue
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Column Widths
            worksheet.column_dimensions['A'].width = 10 # Match
            worksheet.column_dimensions['B'].width = 20 # Event
            worksheet.column_dimensions['C'].width = 20 # Date
            worksheet.column_dimensions['D'].width = 20 # Seller
            worksheet.column_dimensions['E'].width = 15 # Category
            worksheet.column_dimensions['F'].width = 10 # Qty
            worksheet.column_dimensions['G'].width = 12 # Cost
            worksheet.column_dimensions['H'].width = 12 # My Price
            worksheet.column_dimensions['I'].width = 8 # Curr
            worksheet.column_dimensions['J'].width = 50 # Raw
            
            # Data Styles (Alternating colors or logic)
            # Grouping visually by match: alternate background color when match changes?
            
            thin_border = Border(left=Side(style='thin'), 
                                 right=Side(style='thin'), 
                                 top=Side(style='thin'), 
                                 bottom=Side(style='thin'))
            
            current_match = None
            fill_color = 'FFFFFF'
            alt_fill_color = 'F2F2F2' # Light Gray
            current_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                # Check match column (A is 0)
                match_val = row[0].value
                if match_val != current_match:
                    current_match = match_val
                    # Toggle color
                    fill_color = alt_fill_color if fill_color == 'FFFFFF' else 'FFFFFF'
                    current_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                
                for cell in row:
                    cell.fill = current_fill
                    cell.border = thin_border
                    # Align numbers
                    if cell.col_idx in [1, 6, 7, 8]: # Match, Qty, Prices
                         cell.alignment = Alignment(horizontal='center')

        
        output.seek(0)
        
        filename = f"tickets_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        headers = {
            'Content-Disposition': f'attachment; filename="{filename}"'
        }
        
        return StreamingResponse(
            output, 
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers=headers
        )

    except Exception as e:
        print(f'[ERROR] /api/tickets/export: {e}')
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.delete('/api/tickets/offers/{offer_id}')
def delete_offer(offer_id: str):
    """Delete an offer"""
    try:
        from tickets_storage import BASE_DIR, load_offers_from_file, load_index, save_index
        
        index_data = load_index()
        offers_by_id = index_data.get('offers_by_id', {})
        
        if offer_id not in offers_by_id:
             return {'success': False, 'msg': 'Offer not found'} # Idempotent-ish
             
        file_name = offers_by_id[offer_id]['file']
        file_path = BASE_DIR / file_name
        
        # Read and filter
        all_offers = load_offers_from_file(file_path)
        new_offers = [o for o in all_offers if o.get('id') != offer_id]
        
        # Write back
        temp_file = file_path.with_suffix('.jsonl.tmp')
        with open(temp_file, 'w', encoding='utf-8') as f:
            for offer in new_offers:
                f.write(json.dumps(offer, ensure_ascii=False) + '\n')
        temp_file.replace(file_path)
        
        # Update Index (Remove)
        del index_data['offers_by_id'][offer_id]
        
        # Cleanup other indexes? Ideally yes, but complex reverse lookup. 
        # Rebuilding index is safer or acceptable latency. 
        # For now, just removing from primary ID lookup effectively hides it.
        # Ideally trigger a bg index rebuild or do lazy cleanup.
        save_index(index_data)
        
        return {'success': True}
    except Exception as e:
        print(f'[ERROR] delete_offer: {e}')
        raise HTTPException(status_code=500, detail=str(e))

class UpdateOfferPriceRequest(BaseModel):
    lines: List[dict]

@app.put('/api/tickets/offers/{offer_id}')
def update_offer_prices(offer_id: str, request: UpdateOfferPriceRequest):
    """
    Update my_price for lines in an existing offer.
    We need to load the offer, update specific lines, and save it back (overwriting).
    Since we store in weekly files, this is tricky: we must read the weekly file, update, and rewrite.
    """
    try:
        from tickets_storage import BASE_DIR, load_offers_from_file, get_weekly_file_path, update_index_with_offer
        
        # 1. Locate the file containing the offer
        index_data = load_index()
        offers_by_id = index_data.get('offers_by_id', {})
        if offer_id not in offers_by_id:
             raise HTTPException(status_code=404, detail='Offer not found in index')
        
        file_name = offers_by_id[offer_id]['file']
        file_path = BASE_DIR / file_name
        
        # 2. Read all offers from that file
        all_offers = load_offers_from_file(file_path)
        
        target_offer_index = -1
        target_offer = None
        
        for i, offer in enumerate(all_offers):
            if offer.get('id') == offer_id:
                target_offer = offer
                target_offer_index = i
                break
        
        if target_offer_index == -1:
            raise HTTPException(status_code=404, detail='Offer not found in file')
        
        # 3. Update the lines
        # request.lines should contain objects with {_id (index or id), my_price}
        # In current storage, lines don't have stable IDs, only indices.
        # Frontend provides _id which is just index in the array usually.
        
        # Map incoming updates
        updates_map = {str(l.get('_id')): l.get('my_price') for l in request.lines}
        
        current_lines = target_offer.get('lines', [])
        updated_lines = []
        
        for idx, line in enumerate(current_lines):
            # Frontend uses index as _id usually initiated in frontend state
            # We assume the array order is stable.
            # However, careful: frontend generates _id = i. If we rely on index, we must match.
            # Let's assume frontend sends matching indices.
            
            s_idx = str(idx)
            if s_idx in updates_map:
                line['my_price'] = updates_map[s_idx]
            updated_lines.append(line)
            
        target_offer['lines'] = updated_lines
        all_offers[target_offer_index] = target_offer
        
        # 4. Rewrite the file
        # We use a temp file to be safe
        temp_file = file_path.with_suffix('.jsonl.tmp')
        with open(temp_file, 'w', encoding='utf-8') as f:
            for offer in all_offers:
                f.write(json.dumps(offer, ensure_ascii=False) + '\n')
        
        temp_file.replace(file_path)
        
        return {'success': True, 'msg': 'Offer updated'}

    except Exception as e:
        print(f'[ERROR] update_offer_prices: {e}')
        raise HTTPException(status_code=500, detail=str(e))


@app.get('/api/tickets/offers/{offer_id}')
def get_offer(offer_id: str):
    """Return single offer by id"""
    try:
        offer = get_offer_by_id(offer_id)
        if not offer:
            raise HTTPException(status_code=404, detail='Offer not found')
        return offer
    except HTTPException:
        raise
    except Exception as e:
        print(f'[ERROR] /api/tickets/offers/{offer_id}: {e}')
        raise HTTPException(status_code=500, detail=str(e))

@app.post('/api/tickets/index/rebuild')
def rebuild_index_endpoint():
    """Rebuild index by scanning all weekly offers JSONL files"""
    try:
        success = rebuild_index()
        return {'success': success}
    except Exception as e:
        print(f'[ERROR] /api/tickets/index/rebuild: {e}')
        raise HTTPException(status_code=500, detail=str(e))

@app.get('/health')
def health_check():
    """Lightweight health check for Railway"""
    return {'status': 'ok', 'message': 'Server is running'}

# Catch-all route for React SPA - MUST be last (after API routes)
# FastAPI matches routes in order, so specific routes above will be matched first
@app.get('/{full_path:path}')
async def serve_react_app(full_path: str, request: Request):
    """Serve React app for all non-API routes (SPA routing)"""
    # Explicitly exclude API routes and static assets
    # These should never reach here if routes are defined correctly above
    excluded_paths = ['matches', 'history', 'teams', 'health', 'assets', 'vite.svg']
    if any(full_path.startswith(excluded) for excluded in excluded_paths):
        # This shouldn't happen if routes are defined correctly, but just in case
        raise HTTPException(status_code=404, detail="API route not found")
    
    # Serve index.html for all other routes (React Router handles routing)
    index_path = f'{client_dist}/index.html'
    if os.path.exists(index_path):
        return FileResponse(
            index_path,
            media_type='text/html',
            headers={'Cache-Control': 'no-cache'}
        )
    return {'message': 'Build Not Found.'}

if __name__ == '__main__':
    try:
        import time
        startup_time = time.time()
        
        print('\n' + '='*60, flush=True)
        print(f'  [START] VIAGOGO MONITOR - SERVER ONLY (NO SCRAPERS)', flush=True)
        print(f'  [PORT] {PORT}', flush=True)
        print(f'  [DATA] Loading from {DATA_FILE_VIAGOGO} and {DATA_FILE_FTN}', flush=True)
        print('='*60 + '\n', flush=True)
        
        # Quick check - don't load full data at startup, just verify files exist
        print('[INFO] Verifying data files exist...', flush=True)
        if os.path.exists(DATA_FILE_VIAGOGO):
            file_size = os.path.getsize(DATA_FILE_VIAGOGO)
            print(f'[OK] {DATA_FILE_VIAGOGO} exists ({file_size} bytes)', flush=True)
        else:
            print(f'[WARN] {DATA_FILE_VIAGOGO} not found', flush=True)
        
        if os.path.exists(DATA_FILE_FTN):
            file_size = os.path.getsize(DATA_FILE_FTN)
            print(f'[OK] {DATA_FILE_FTN} exists ({file_size} bytes)', flush=True)
        else:
            print(f'[WARN] {DATA_FILE_FTN} not found', flush=True)
        
        # Verify frontend build exists
        print('[INFO] Verifying frontend build...', flush=True)
        if not os.path.exists(client_dist):
            print(f'[ERROR] Frontend dist directory not found: {client_dist}', flush=True)
        elif not os.path.exists(f'{client_dist}/index.html'):
            print(f'[ERROR] Frontend index.html not found at {client_dist}/index.html', flush=True)
        else:
            print(f'[OK] Frontend build verified: {client_dist}/index.html', flush=True)
        
        elapsed = time.time() - startup_time
        print(f'[INFO] Startup checks completed in {elapsed:.2f}s', flush=True)
        
        print(f'[INFO] Starting server on 0.0.0.0:{PORT}...', flush=True)
        print(f'[INFO] Railway PORT env: {os.environ.get("PORT", "NOT SET")}', flush=True)
        print('[INFO] Server starting now...', flush=True)
        
        # Start uvicorn server - this blocks until server stops
        uvicorn.run(
            app, 
            host='0.0.0.0', 
            port=PORT, 
            log_level='info'
        )
    except KeyboardInterrupt:
        print('\n[INFO] Server stopped by user', flush=True)
    except Exception as e:
        print(f'[FATAL] Server startup failed: {e}', flush=True)
        import traceback
        traceback.print_exc()
        sys.exit(1)

